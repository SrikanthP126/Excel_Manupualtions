from openpyxl import load_workbook

# Load the disposition workbook
disposition_file = 'path_to_your_disposition_file.xlsx'  # Replace with your file path
disposition_wb = load_workbook(disposition_file)

# Load the deleted objects workbook
deleted_objects_file = 'path_to_your_deleted_objects_file.xlsx'  # Replace with your file path
deleted_objects_wb = load_workbook(deleted_objects_file)

# For each sheet in the disposition workbook, find and remove deleted objects
for sheet_name in disposition_wb.sheetnames:
    disposition_sheet = disposition_wb[sheet_name]
    deleted_objects_sheet = deleted_objects_wb[sheet_name]
    
    # Assuming 'Object Name' is in the first column (A) and starts from the second row
    deleted_objects = set(deleted_objects_sheet['A'][1:])
    
    rows_to_delete = []
    for row in disposition_sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        object_name = row[0]
        if object_name in deleted_objects:
            rows_to_delete.append(row)
    
    # Delete the rows from the disposition sheet
    for row in rows_to_delete:
        disposition_sheet.delete_rows(row[0].row)

# Save the cleaned disposition workbook
disposition_wb.save('cleaned_disposition_file.xlsx')

print("Process completed. Check the 'cleaned_disposition_file.xlsx' for the results.")
